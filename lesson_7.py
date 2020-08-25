import openpyxl
import requests

# 读取用例并封装
def read_data(filename, sheetname):
    excel = openpyxl.load_workbook(filename)
    sheet = excel[sheetname]
    max_row = sheet.max_row
    case_list = []
    for i in range(2, max_row+1):
        dict1 = dict(
            id=sheet.cell(row=i, column=1).value,
            url=sheet.cell(row=i, column=5).value,
            data=sheet.cell(row=i, column=6).value,
            expected=sheet.cell(row=i, column=7).value)
        case_list.append(dict1)
    return case_list

#发送请求

#登录接口
def api_log_func(url,body):
    log_url = url
    log_header = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    log_body = body
    log_res = requests.post(url=log_url,json=log_body,headers=log_header)
    result = log_res.json()
    return result


#写入数据
def write_result(filename, sheetname, row, column, final_result):
    excel = openpyxl.load_workbook(filename)
    sheet = excel[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    excel.save(filename)

# 取出具体用例数据
def execute_func(filename,sheetname):
    res = read_data(filename, sheetname)  #用res接收调用read_data方法取出来的数据
    for testcase in res:                                 #用testcase变量来遍历测试用例文件
        test_id = testcase.get('id')                      #取出id
        test_url = testcase.get('url')                      #取出url
        test_data = testcase.get('data')                      #取出data,从excel取出来的数据类型都是“str”类型的
        test_data = eval(test_data)                             #用eval()函数将引号去掉
        test_expected = testcase.get('expected')                 #取出expected
        test_expected = eval(test_expected)
        expected_msg = test_expected.get('msg')
        # print(test_id, test_url, test_data, test_expected)
        res_1 = api_log_func(url=test_url,body=test_data)
        # print(res_1)
        real_msg = res_1.get('msg')
        print('预期结果为：{}'.format(expected_msg))
        print('实际结果为：{}'.format(real_msg))
        if expected_msg == real_msg:
            print('此条用例通过!')
            final_result = 'PASS'
        else:
            print('此条用例不通过！！！')
            final_result = 'NG'
        print('*'*100)
        write_result(filename,sheetname,test_id+1,8,final_result)

execute_func('test_case_api.xlsx','register')
execute_func('test_case_api.xlsx','login')

